import logging

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from mulaco.core.app import App
from mulaco.excel.utils import excel_col_alpha2num
from mulaco.models.bo_model import ExcelSheetBO
from mulaco.models.dto_model import ExcelDTO, SheetDTO
from mulaco.models.mapper import exsh_bo_map_po
from mulaco.models.po_model import CellInfoPO

log = logging.getLogger(__name__)


class ExcelLoader:
    CACHE_TBL = "excels"
    # -------------------- loader --------------------

    def __init__(self, app: App):
        self.db = app.db
        self.cache = app.cache

    def load_excel(self, excel: ExcelDTO):
        """
        TODO: 优化建议
        在初始化的时候
        把要的行和列读到缓存，再进行持久化操作，
        否则数据量大的话一张表就要 1-2 分钟，很费时。
        """
        ex_name = excel.excel_name
        try:
            log.info(f"开始加载 {ex_name} ...")
            wb = load_workbook(excel.src_path)
            for sheet_dto in excel.sheets:
                sheet = wb[sheet_dto.sheet_name]
                self._set_db_exsh_meta(ex_name, sheet_dto)
                self._set_cache_exsh_meta_data(
                    sheet.max_row, sheet.max_column, excel, sheet_dto
                )
                self._set_db_sheet_raw_data(ex_name, sheet, sheet_dto)
            log.info(f"完成加载 {ex_name} ...")
        except Exception:
            # log.exception(e)
            log.error(f"加载出错!!! {ex_name} ！")
        finally:
            wb.close()

    def _set_db_exsh_meta(self, excel_name: str, sheet_dto: SheetDTO) -> int:
        """持久化存 ExcelSheet 元数据"""
        exsh_bo = ExcelSheetBO(
            excel=excel_name,
            sheet=sheet_dto.sheet_name,
            header=sheet_dto.header_row,
        )
        exsh_po = exsh_bo_map_po(exsh_bo)
        return self.db.upsert_exsh(exsh_po)

    def _set_cache_exsh_meta_data(
        self, max_row: int, max_col: int, excel_dto: ExcelDTO, sheet_dto: SheetDTO
    ):
        """缓存元数据"""
        sheet_dto.max_row = max_row
        sheet_dto.max_col = max_col
        _d = excel_dto.to_dict()
        self.cache.set(excel_dto.excel_name, _d, self.CACHE_TBL)

    # TODO 优化函数
    # def cache_worksheet_by_cols(
    #     self, sheet: Worksheet, lang_cols: dict[str, list[int]]
    # ):
    #     for lang, cols in lang_cols.items():
    #         pass

    def _set_db_sheet_raw_data(
        self, ex_name: str, sheet: Worksheet, sheet_dto: SheetDTO
    ):
        """持久化 Sheet 中的原始数据"""
        # 存 EXSH 中的数据
        max_row = sheet_dto.max_row
        sh_name = sheet_dto.sheet_name
        exsh_po = self.db.get_exsh_by_name(ex_name, sh_name)
        # 遍历每个 languages 对象，一般来说是 zh 和 en
        for src_lang, col_list in sheet_dto.lang_cols.items():
            # 遍历每个 col
            for col_alpha in col_list:
                col = excel_col_alpha2num(col_alpha)
                # 遍历每行
                for row in range(sheet_dto.header_row + 1, max_row + 1):
                    raw_text = sheet.cell(row, col).value
                    cell_po = CellInfoPO(
                        exsh_id=exsh_po.id,
                        row=row,
                        col=col,
                        src_lang=src_lang,
                        raw_text=raw_text,
                    )
                    self.db.upsert_cell(cell_po)
                log.debug(f"加载 {ex_name}.{sh_name} 的 {col_alpha} 列")
