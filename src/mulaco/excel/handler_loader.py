import logging

import xlwings as xw
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from mulaco.base.db import JsonCache
from mulaco.core.db_models import CellInfo, ExcelSheet
from mulaco.core.service import DbService
from mulaco.excel.model import ExcelDTO, SheetDTO
from mulaco.excel.utils import excel_col_alpha2num

log = logging.getLogger(__name__)


class ExcelHandler:
    cache_tbl = "excels"

    def __init__(self, excel: ExcelDTO, db: DbService, cache: JsonCache):
        self.excel = excel
        self.db = db
        self.cache = cache


class ExcelLoader(ExcelHandler):

    def loader(self):
        if self.excel.skip:
            log.debug(f"{self.excel} 跳过")
            return
        try:
            wb = load_workbook(self.excel.src_path)
            for sheet_dto in self.excel.sheets:
                sheet = wb[sheet_dto.sheet_name]
                self.persist_exsh_meta(sheet_dto)
                self.cache_exsh_meta_data(sheet.max_row, sheet.max_column, sheet_dto)
                self.persist_sheet_raw_data(sheet, sheet_dto)
        except Exception as e:
            log.exception(e)
            log.error(f"{self.excel} 载入数据时发生错误")
        finally:
            wb.close()

    def persist_exsh_meta(self, sheet_dto: SheetDTO) -> int:
        """持久化存 ExcelSheet 元数据"""
        exsh_po = ExcelSheet(
            excel=self.excel.excel_name,
            sheet=sheet_dto.sheet_name,
            header=sheet_dto.header_row,
        )
        exsh = self.db.upsert_exsh(exsh_po)
        return exsh

    def cache_exsh_meta_data(self, max_row: int, max_col: int, sheet_dto: SheetDTO):
        """缓存元数据"""
        sheet_dto.max_row = max_row
        sheet_dto.max_col = max_col
        _d = self.excel.to_dict()
        self.cache.set(self.excel.excel_name, _d, self.cache_tbl)

    def persist_sheet_raw_data(self, sheet: Worksheet, sheet_dto: SheetDTO):
        """持久化 Sheet 中的原始数据"""
        max_row = sheet_dto.max_row
        col_dto = sheet_dto.lang_cols
        ex_name = self.excel.excel_name
        sh_name = sheet_dto.sheet_name
        # 遍历每个 languages 对象，一般来说是 zh 和 en
        for col_dto in sheet_dto.lang_cols:
            lang = col_dto.src_lang
            # 遍历每个 col
            for col_alpha in col_dto.cols:
                col = excel_col_alpha2num(col_alpha)
                # 遍历每行
                for row in range(sheet_dto.header_row + 1, max_row + 1):
                    loc = f"{col_alpha}{row}"
                    raw_text = sheet[loc].value
                    cell = CellInfo(row=row, col=col, src_lang=lang, raw_text=raw_text)
                    self.db.add_cell_info(ex_name, sh_name, cell)


class ExcelExporter(ExcelHandler):

    def writer(self): ...
