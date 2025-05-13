import logging
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from mulaco.core.app import App
from mulaco.excel.utils import excel_col_alpha2num
from mulaco.fix.parser import CellParser
from mulaco.models.bo_model import ExcelSheetBO
from mulaco.models.dto_model import ExcelDTO, SheetDTO

log = logging.getLogger(__name__)


class ExcelExporter:
    CACHE_EXCEL_TBL = "excels"

    # TODO 这里的参数是传 excel_name 还是传 excel_dto 呢
    def __init__(self, app: App):
        self.db = app.db
        self.cache = app.cache
        self.dst_langs = app.dst_langs
        self.langs_mapper = app.langs_mapper

    # -------------------- export --------------------
    def export_excel(self, excel: ExcelDTO):
        excel = self.flush_excel(excel)
        try:
            # 第一步，复制数据
            self._copy_excel(excel)
            wb = load_workbook(excel.dst_path)
            for sheet_dto in excel.sheets:
                sheet = wb[sheet_dto.sheet_name]
                # 该参数写死，但是之后是可以配置的
                self.write_sheet(excel, sheet, sheet_dto, "en")

            # 保存
            wb.save(excel.dst_path)
        except Exception:
            log.error(f"{excel.excel_name} 写入数据时发送错误")
        finally:
            wb.close()

    def flush_excel(self, excel_dto: ExcelDTO) -> ExcelDTO:
        d = self.cache.get(excel_dto.excel_name, self.CACHE_EXCEL_TBL)
        return ExcelDTO.from_dict(d)

    def write_sheet(
        self, ex_dto: ExcelDTO, sheet: Worksheet, sh_dto: SheetDTO, src: str
    ):
        """TODO 优化建议，直接将准备好的数据批量写入"""
        # db 选择行
        ex_name = ex_dto.excel_name
        sh_name = sh_dto.sheet_name
        max_col = sh_dto.max_col
        header_row = sh_dto.header_row
        #
        exsh_bo = ExcelSheetBO(excel=ex_name, sheet=sh_name, header=None)
        cols = sh_dto.lang_cols[src]
        total_dst_lang = len(self.dst_langs)
        for idx, src_col in enumerate(cols):
            # 主要是标记
            # idx 主要是标记序号的
            col_num = excel_col_alpha2num(src_col)
            for dst in self.dst_langs:
                # TODO 从 Lang 模型中提取 offset 或者，可以在设置 lang 时设置
                offset = self.langs_mapper[dst].order
                res = self.db.get_all_write_trans(exsh_bo, src, dst, col_num)
                # 计算列
                d_col = max_col + offset + total_dst_lang * idx
                # 写入表头
                header_text = src_col + "列" + self.langs_mapper[dst].name
                sheet.cell(header_row, d_col).value = header_text

                for cell_po, trans_po in res:
                    c_row = cell_po.row
                    text = trans_po.write_text
                    sheet.cell(c_row, d_col).value = text
                log.debug(f"完成 {ex_name}.{sh_name} lang={dst} 中的写入")
        # 修复 col
        # 该步骤要删除
        # if ex_dto.refs:
        #     self.fix_sheet_ref(ex_dto, sh_dto, sheet)

    # TODO 该函数要删除
    def fix_sheet_ref(self, ex_dto: ExcelDTO, sh_dto: SheetDTO, sheet: Worksheet):
        """修复所有表的 ref"""
        max_row = sh_dto.max_row
        # 注册一个 Parser
        parser = CellParser()
        for cols in sh_dto.lang_cols.values():
            for col_alpha in cols:
                col_num = excel_col_alpha2num(col_alpha)
                for row in range(1, max_row + 1):
                    # 如果存在
                    text = sheet.cell(row, col_num).value
                    dir_name = str(Path(ex_dto.dst_path).parent.absolute())
                    dir_name = dir_name.replace("\\", "/")
                    t = parser.self_fix_ref(dir_name, ex_dto.excel_name, text)
                    sheet.cell(row, col_num).value = t
        log.debug(f"已修复 {ex_dto.excel_name}.{sh_dto.sheet_name} 中的链接")

    def _copy_excel(self, excel: ExcelDTO):
        """复制文件到指定位置，如果存在，直接返回"""
        src = Path(excel.src_path)
        dst = Path(excel.dst_path)
        if dst.exists():
            return
        else:
            shutil.copy2(src, dst)
